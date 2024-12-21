#!/usr/bin/env python3
"""
Enhanced 3GPP PM Parser
Output handler implementations for different output formats.
"""

import asyncio
import csv
import logging
import sqlite3
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger('PMParser.Output')

class OutputHandler(ABC):
    """Abstract base class for output handlers"""
    
    @abstractmethod
    async def write(self, data: List[Dict]):
        """Write data to output"""
        pass

class ExcelOutputHandler(OutputHandler):
    """Handler for Excel output"""
    
    def __init__(self, filename: str = None):
        self.filename = filename or f"pm_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    async def write(self, data: List[Dict]):
        """Write data to Excel file"""
        if not data:
            logger.warning("No data to write to Excel")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "PM Data"

        # Write headers
        headers = ['Time', 'MeasInfoId', 'MeasObjLdn', 'P', 'Value', 'MeasType']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row, item in enumerate(data, 2):
            ws.cell(row=row, column=1).value = item['endTime']
            ws.cell(row=row, column=2).value = item['measInfoId']
            ws.cell(row=row, column=3).value = item['measObjLdn']
            ws.cell(row=row, column=4).value = item['p']
            ws.cell(row=row, column=5).value = item['value']
            ws.cell(row=row, column=6).value = item['measType']

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(self.filename)
        logger.info(f"Data written to Excel file: {self.filename}")

class SQLiteOutputHandler(OutputHandler):
    """Handler for SQLite output"""
    
    def __init__(self, db_file: str = None):
        self.db_file = db_file or f"pm_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        self._create_tables()

    def _create_tables(self):
        """Create necessary database tables"""
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS measData (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    endTime TIMESTAMP,
                    measInfoId VARCHAR,
                    measObjLdn VARCHAR,
                    p INTEGER,
                    value FLOAT,
                    measType VARCHAR,
                    UNIQUE(endTime, measInfoId, measObjLdn, p)
                )
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_measData_time ON measData(endTime)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_measData_info ON measData(measInfoId)")

    async def write(self, data: List[Dict]):
        """Write data to SQLite database"""
        if not data:
            logger.warning("No data to write to SQLite")
            return

        with sqlite3.connect(self.db_file) as conn:
            cursor = conn.cursor()
            
            # Prepare data for batch insert
            insert_data = [
                (
                    item['endTime'],
                    item['measInfoId'],
                    item['measObjLdn'],
                    item['p'],
                    item['value'],
                    item['measType']
                )
                for item in data
            ]
            
            # Batch insert with error handling
            try:
                cursor.executemany(
                    """
                    INSERT OR REPLACE INTO measData 
                    (endTime, measInfoId, measObjLdn, p, value, measType)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    insert_data
                )
                conn.commit()
                logger.info(f"Data written to SQLite database: {self.db_file}")
            except sqlite3.Error as e:
                logger.error(f"Error writing to SQLite: {str(e)}")
                raise

class CSVOutputHandler(OutputHandler):
    """Handler for CSV output"""
    
    def __init__(self, filename: str = None):
        self.filename = filename or f"pm_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    async def write(self, data: List[Dict]):
        """Write data to CSV file"""
        if not data:
            logger.warning("No data to write to CSV")
            return

        fieldnames = ['endTime', 'measInfoId', 'measObjLdn', 'p', 'value', 'measType']
        
        with open(self.filename, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
            
        logger.info(f"Data written to CSV file: {self.filename}")

class OutputHandlerFactory:
    """Factory class for creating output handlers"""
    
    @staticmethod
    def create(output_type: str, **kwargs) -> OutputHandler:
        """Create an output handler based on type"""
        handlers = {
            'excel': ExcelOutputHandler,
            'sqlite': SQLiteOutputHandler,
            'csv': CSVOutputHandler
        }
        
        handler_class = handlers.get(output_type.lower())
        if not handler_class:
            raise ValueError(f"Unsupported output type: {output_type}")
            
        return handler_class(**kwargs)